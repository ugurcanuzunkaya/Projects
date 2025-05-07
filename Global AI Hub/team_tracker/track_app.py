import os
import streamlit as st
import pandas as pd
from datetime import date
from openpyxl import Workbook, load_workbook

# Determine the folder where this script is located
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
# Save the Excel file in the same directory as the script
FILE_PATH = os.path.join(SCRIPT_DIR, "team_meetings.xlsx")

def initialize_workbook(file_path):
    """
    If the Excel file doesn't exist, create it and add the header row.
    """
    if not os.path.exists(file_path):
        wb = Workbook()
        ws = wb.active
        ws.title = "Meetings"
        # Create header row with additional columns for absent members with/without reason
        ws.append(["Date", "Location", "Attendees", "Absent Members (No Reason)", "Absent Members (With Reason)"])
        wb.save(file_path)
        
# Initialize the Excel workbook at startup
initialize_workbook(FILE_PATH)

def add_meeting(file_path, meeting_date, location, attendees, absent_no_reason, absent_with_reason):
    """
    Append a new meeting entry to the Excel file.
    """
    wb = load_workbook(file_path)
    ws = wb.active
    # Append the meeting data; convert the date to string
    ws.append([str(meeting_date), location, attendees, absent_no_reason, absent_with_reason])
    wb.save(file_path)

# Streamlit UI
st.title("Team Meetings Tracker")
st.write("Enter the meeting details below:")

with st.form(key="meeting_form"):
    # Meeting Date: A date picker with today's date as default
    meeting_date = st.date_input("Meeting Date", date.today())
    
    # Meeting Location: A text input field
    location = st.text_input("Meeting Location")
    
    # Attendees: Enter as comma-separated values
    attendees = st.text_input("Attendees (comma separated)")
    
    # Absent Members (No Reason): Enter names as comma-separated values
    absent_no_reason = st.text_input("Absent Members (no reason provided, comma separated)")
    
    # Absent Members (With Reason): Enter in the format Name: Reason, comma separated
    absent_with_reason = st.text_input("Absent Members (with reason provided, format: Name: Reason, comma separated)")
    
    # Submit button for the form
    submit_button = st.form_submit_button(label="Submit Meeting")

# When the form is submitted, add the meeting to the Excel file
if submit_button:
    add_meeting(FILE_PATH, meeting_date, location, attendees, absent_no_reason, absent_with_reason)
    st.success(f"Meeting on {meeting_date} saved successfully!")

# Optionally, display the current meetings data from the Excel file
if st.checkbox("Show all meetings"):
    try:
        df = pd.read_excel(FILE_PATH)
        st.dataframe(df)
    except Exception as e:
        st.error(f"Error loading data: {e}")
